print(7%(5//2))



print("Hello World")

usertext = input("What is your name? ")
print("Hello", usertext)

num1 = input('Enter first number: ')
num2 = input('Enter second number: ')

sum = float(num1) + float(num2)

print('The sum of {0} and {1} is {2}'.format(num1, num2, sum))

num1 = input('Enter first number: ')
num2 = input('Enter second number: ')

average =(int(num1) + int(num2))

print('average:{0} '.format(average))

visagrade = input('enter your visa grade : ')
finalgrade = input('enter your final grade : ')
average =(float(visagrade)*0.3)+(float(finalgrade)*0.7)
print("average :{0} ".format(average))

firstexam = input('your first exam : ')
secondexam = input('your second exam : ')
thirdexam = input('your third exam : ')
average =(float(firstexam)+float(secondexam)+float(thirdexam))/3
print("average :{0} ".format(average))

num = int(input("Enter a number: "))
if (num % 2) == 0:
   print("{0} is Even".format(num))
else:
   print("{0} is Odd".format(num))

   num = float(input("Enter a number: "))
   if num > 0:
       print("Positive number")
   elif num == 0:
       print("Zero")
   else:
       print("Negative number")

       print("body mass index calculation program")
       height = float(input("enter height (m):"))
       weight = int(input("enter weight (kg):"))

       index = weight / (height * height)

       if index <= 18:
           print("\n underweight BMİ:{}".format(index))
       elif index > 18 and index <= 25:
           print("\n overweight BMİ:{}".format(index))
       elif index > 25 and index <= 30:
           print("\n obese BMİ:{}".format(index))
       elif index > 30:
           print("\n severely obese BMİ:{}".format(index))


           age = input('enter age : ')
           if (int(age) < 18):
               print("Your Age Is Not Eligible To Get A Driver's License")
           else:
               print("Your Age Is Eligible To Get Your License")

               for i in range(1, 101):
                   print(i)

                   for i in range(1, 101):
                       if i % 2 == 0:
                           print(i)

                           for i in range(1, 101):
                               if i % 2 != 0:
                                   print(i)

                                   for i in range(1, 101):
                                       if i % 3 == 0 or i % 5 == 0:
                                           print(i)

                                           num = input('enter number : ')
                                           for i in range(1, int(num) + 1):
                                               print(i)

                                               short = input('Enter short side : ')
                                               tall = input('Enter tall side : ')
                                               area = int(short) * int(tall)
                                               perimeter = 2 * (int(short) + int(tall))
                                               print("area: {0}".format(alan))
                                               print("perimeter: {0}".format(cevre))

                                               word = 'mrhuseyin'
                                               for char in word:
                                                   print(char)

                                                   sumofnumbers = 0;
                                                   num1 = input('first number: ')
                                                   num2 = input('second number: ')
                                                   for i in range(int(sayi1) + 1, int(sayi2)):
                                                       sumofnumbers += i
                                                   print("Sum of numbers between {0} and {1} : {2}".format(num1, num2,
                                                                                                           sumofnumbers))

                                                   selection = input("Press (1) for Cinema, (2) for Theater : ")
                                                   student = input("Are you student(Y/N) : ")
                                                   price = 0
                                                   # non-discounted fee calculation
                                                   if selection == '1':
                                                       price = 10  # cinema
                                                   elif selection == '2':
                                                       price = 5  # theatre
                                                   # student discount
                                                   if student == 'Y' or student == 'y':
                                                       price = price / 2  # %50
                                                   print(" The fee you have to pay :{}".format(price))

                                                   num = int(input("Enter a number: "))

                                                   if num > 1:
                                                       for i in range(2, num):
                                                           if (num % i) == 0:
                                                               print(num, "is not a prime number")
                                                               print(i, "times", num // i, "is", num)
                                                               break
                                                       else:
                                                           print(num, "is a prime number")

                                                   else:
                                                       print(num, "is not a prime number")

                                                       newsalary = 0
                                                       salary = input("enter new salary : ")
                                                       raise = input("salary raise rate(%) : ")
                                                       newsalary = int(salary) + (int(salary) * int(
                                                       raise) / 100)
                                                       print("increased salary :", newsalary)

                                                       import math


                                                       def find_Diameter(radius):
                                                           return 2 * radius


                                                       def find_Circumference(radius):
                                                           return 2 * math.pi * radius


                                                       def find_Area(radius):
                                                           return math.pi * radius * radius


                                                       r = float(input(' Please Enter the radius of a circle: '))

                                                       diameter = find_Diameter(r)
                                                       circumference = find_Circumference(r)
                                                       area = find_Area(r)

                                                       print("\n Diameter Of a Circle = %.2f" % diameter)
                                                       print(" Circumference Of a Circle = %.2f" % circumference)
                                                       print(" Area Of a Circle = %.2f" % area)


                                                       def areaRectangle(a, b):
                                                           return (a * b)


                                                       def perimeterRectangle(a, b):
                                                           return (2 * (a + b))


                                                       a = 5;
                                                       b = 6;
                                                       print("Area = ", areaRectangle(a, b))

                                                       print("Perimeter = ", perimeterRectangle(a, b))

                                                       import random
                                                       import math

                                                       # Taking Inputs
                                                       lower = int(input("Enter Lower bound:- "))

                                                       # Taking Inputs
                                                       upper = int(input("Enter Upper bound:- "))

                                                       # generating random number between
                                                       # the lower and upper
                                                       x = random.randint(lower, upper)
                                                       print("\n\tYou've only ",
                                                             round(math.log(upper - lower + 1, 2)),
                                                             " chances to guess the integer!\n")

                                                       # Initializing the number of guesses.
                                                       count = 0

                                                       # for calculation of minimum number of
                                                       # guesses depends upon range
                                                       while count < math.log(upper - lower + 1, 2):
                                                           count += 1

                                                           # taking guessing number as input
                                                           guess = int(input("Guess a number:- "))

                                                           # Condition testing
                                                           if x == guess:
                                                               print("Congratulations you did it in ",
                                                                     count, " try")
                                                               # Once guessed, loop will break
                                                               break
                                                           elif x > guess:
                                                               print("You guessed too small!")
                                                           elif x < guess:
                                                               print("You Guessed too high!")

                                                       # If Guessing is more than required guesses,
                                                       # shows this output.
                                                       if count >= math.log(upper - lower + 1, 2):
                                                           print("\nThe number is %d" % x)
                                                           print("\tBetter Luck Next time!")

                                                           import datetime

                                                           date = str(input('Enter the date(for example:09 02 2019):'))
                                                           day_name = ['Monday', 'Tuesday', 'Wednesday', 'Thursday',
                                                                       'Friday', 'Saturday', 'Sunday']
                                                           day = datetime.datetime.strptime(date, '%d %m %Y').weekday()
                                                           print(day_name[day])


                                                           def find_missing(lst):
                                                               return [x for x in range(lst[0], lst[-1] + 1)
                                                                       if x not in lst]

                                                               # Driver code


                                                           lst = [1, 2, 4, 6, 7, 9, 10]
                                                           print(find_missing(lst))
                                                           Output:
                                                           [3, 5, 8]

                                                           char_list = ["a", "b", "c"]
                                                           string = "abcd"
                                                           matched_list = [characters in char_list for characters in
                                                                           string]
                                                           print(matched_list)
                                                           OUTPUT
                                                           [True, True, True, False]
                                                           string_contains_chars = all(matched_list)
                                                           print(string_contains_chars)

                                                           total = 0
                                                           evenSums = 0
                                                           oddSums = 0
                                                           done = False
                                                           while (not done):
                                                               user_in = input(
                                                                   "Give me an integer or type 'done' to be done.")
                                                               if (user_in.lower() == "done"):
                                                                   done = True
                                                               else:
                                                                   # assuming they've typed in an integer
                                                                   total += int(user_in)
                                                                   if user_in % 2 == 0:
                                                                       evenSums += user_in
                                                                       evenAverage = evenSums / user_in
                                                                   else:
                                                                       oddSums += user_in
                                                                       oddAverage = oddSums / user_in
                                                           print(total)
                                                           print("Even Average: " + str(evenAverage))
                                                           print("Odd Average: " + str(oddAverage))

                                                           import pymysql.cursors

                                                           # Database connection sentence
                                                           connection = pymysql.connect(host='localhost',
                                                                                        user='root',
                                                                                        password='',
                                                                                        db='students',
                                                                                        charset='utf8mb4',
                                                                                        cursorclass=pymysql.cursors.DictCursor)
                                                           try:
                                                               with connection.cursor() as cursor:
                                                                   # single line reading
                                                                   sql = "SELECT `id`, `firstname`,`lastname` FROM `users`"
                                                                   cursor.execute(sql)
                                                                   for row in cursor.fetchall():
                                                                       # read all lines
                                                                       firstname = str(row["firstname"])
                                                                       lastname = str(row["lastname"])
                                                                       # print on screen
                                                                       print("firstname : " + firstname)
                                                                       print("lastname : " + lastname)
                                                       finally:
                                                       connection.close()

                                                       from tkinter import *

                                                       window = Tk()
                                                       # add widgets here

                                                       window.title('Hello Python')
                                                       window.geometry("300x200+10+20")
                                                       window.mainloop()

                                                       from tkinter import *
                                                       from tkinter import messagebox

                                                       window = Tk()
                                                       window.title("mrhuseyin.medium.com")
                                                       window.geometry("400x300")
                                                       # plotting grid forms
                                                       application = Frame(window)
                                                       application.grid()
                                                       L1 = Label(uygulama, text="Enter your name")
                                                       L1.grid(padx=110, pady=10)
                                                       E1 = Entry(application, bd=2)
                                                       E1.grid(padx=110, pady=3)
                                                       # draw form
                                                       window.mainloop()

                                                       from tkinter import *

                                                       from tkinter import messagebox

                                                       window = Tk()

                                                       window.title("mrhuseyin.medium.com")
                                                       window.geometry("400x300")

                                                       # grid form çizdirme
                                                       application = Frame(window)
                                                       application.grid()

                                                       Lb1 = Listbox(application)
                                                       Lb1.insert(1, "Python")
                                                       Lb1.insert(2, "C#")
                                                       Lb1.insert(3, "JAVA")
                                                       Lb1.insert(4, "JAVASCRIPT")
                                                       Lb1.grid(padx=110, pady=10)

                                                       # draw form
                                                       pencere.mainloop()

                                                       # import openpyxl and tkinter modules
                                                       from openpyxl import *
                                                       from tkinter import *

                                                       # globally declare wb and sheet variable

                                                       # opening the existing excel file
                                                       wb = load_workbook('C:\\Users\\Admin\\Desktop\\excel.xlsx')

                                                       # create the sheet object
                                                       sheet = wb.active


                                                       def excel():

                                                           # resize the width of columns in
                                                           # excel spreadsheet
                                                           sheet.column_dimensions['A'].width = 30
                                                           sheet.column_dimensions['B'].width = 10
                                                           sheet.column_dimensions['C'].width = 10
                                                           sheet.column_dimensions['D'].width = 20
                                                           sheet.column_dimensions['E'].width = 20
                                                           sheet.column_dimensions['F'].width = 40
                                                           sheet.column_dimensions['G'].width = 50

                                                           # write given data to an excel spreadsheet
                                                           # at particular location
                                                           sheet.cell(row=1, column=1).value = "Name"
                                                           sheet.cell(row=1, column=2).value = "Course"
                                                           sheet.cell(row=1, column=3).value = "Semester"
                                                           sheet.cell(row=1, column=4).value = "Form Number"
                                                           sheet.cell(row=1, column=5).value = "Contact Number"
                                                           sheet.cell(row=1, column=6).value = "Email id"
                                                           sheet.cell(row=1, column=7).value = "Address"


                                                       # Function to set focus (cursor)
                                                       def focus1(event):
                                                           # set focus on the course_field box
                                                           course_field.focus_set()

                                                           # Function to set focus


                                                       def focus2(event):
                                                           # set focus on the sem_field box
                                                           sem_field.focus_set()

                                                           # Function to set focus


                                                       def focus3(event):
                                                           # set focus on the form_no_field box
                                                           form_no_field.focus_set()

                                                           # Function to set focus


                                                       def focus4(event):
                                                           # set focus on the contact_no_field box
                                                           contact_no_field.focus_set()

                                                           # Function to set focus


                                                       def focus5(event):
                                                           # set focus on the email_id_field box
                                                           email_id_field.focus_set()

                                                           # Function to set focus


                                                       def focus6(event):
                                                           # set focus on the address_field box
                                                           address_field.focus_set()

                                                           # Function for clearing the


                                                       # contents of text entry boxes
                                                       def clear():

                                                           # clear the content of text entry box
                                                           name_field.delete(0, END)
                                                           course_field.delete(0, END)
                                                           sem_field.delete(0, END)
                                                           form_no_field.delete(0, END)
                                                           contact_no_field.delete(0, END)
                                                           email_id_field.delete(0, END)
                                                           address_field.delete(0, END)

                                                           # Function to take data from GUI


                                                       # window and write to an excel file
                                                       def insert():

                                                           # if user not fill any entry
                                                           # then print "empty input"
                                                           if (name_field.get() == "" and
                                                                   course_field.get() == "" and
                                                                   sem_field.get() == "" and
                                                                   form_no_field.get() == "" and
                                                                   contact_no_field.get() == "" and
                                                                   email_id_field.get() == "" and
                                                                   address_field.get() == ""):

                                                               print("empty input")

                                                           else:

                                                               # assigning the max row and max column
                                                               # value upto which data is written
                                                               # in an excel sheet to the variable
                                                               current_row = sheet.max_row
                                                               current_column = sheet.max_column

                                                               # get method returns current text
                                                               # as string which we write into
                                                               # excel spreadsheet at particular location
                                                               sheet.cell(row=current_row + 1,
                                                                          column=1).value = name_field.get()
                                                               sheet.cell(row=current_row + 1,
                                                                          column=2).value = course_field.get()
                                                               sheet.cell(row=current_row + 1,
                                                                          column=3).value = sem_field.get()
                                                               sheet.cell(row=current_row + 1,
                                                                          column=4).value = form_no_field.get()
                                                               sheet.cell(row=current_row + 1,
                                                                          column=5).value = contact_no_field.get()
                                                               sheet.cell(row=current_row + 1,
                                                                          column=6).value = email_id_field.get()
                                                               sheet.cell(row=current_row + 1,
                                                                          column=7).value = address_field.get()

                                                               # save the file
                                                               wb.save('C:\\Users\\Admin\\Desktop\\excel.xlsx')

                                                               # set focus on the name_field box
                                                               name_field.focus_set()

                                                               # call the clear() function
                                                               clear()

                                                               # Driver code


                                                       if __name__ == "__main__":
                                                           # create a GUI window
                                                           root = Tk()

                                                           # set the background colour of GUI window
                                                           root.configure(background='light green')

                                                           # set the title of GUI window
                                                           root.title("registration form")

                                                           # set the configuration of GUI window
                                                           root.geometry("500x300")

                                                           excel()

                                                           # create a Form label
                                                           heading = Label(root, text="Form", bg="light green")

                                                           # create a Name label
                                                           name = Label(root, text="Name", bg="light green")

                                                           # create a Course label
                                                           course = Label(root, text="Course", bg="light green")

                                                           # create a Semester label
                                                           sem = Label(root, text="Semester", bg="light green")

                                                           # create a Form No. lable
                                                           form_no = Label(root, text="Form No.", bg="light green")

                                                           # create a Contact No. label
                                                           contact_no = Label(root, text="Contact No.",
                                                                              bg="light green")

                                                           # create a Email id label
                                                           email_id = Label(root, text="Email id", bg="light green")

                                                           # create a address label
                                                           address = Label(root, text="Address", bg="light green")

                                                           # grid method is used for placing
                                                           # the widgets at respective positions
                                                           # in table like structure .
                                                           heading.grid(row=0, column=1)
                                                           name.grid(row=1, column=0)
                                                           course.grid(row=2, column=0)
                                                           sem.grid(row=3, column=0)
                                                           form_no.grid(row=4, column=0)
                                                           contact_no.grid(row=5, column=0)
                                                           email_id.grid(row=6, column=0)
                                                           address.grid(row=7, column=0)

                                                           # create a text entry box
                                                           # for typing the information
                                                           name_field = Entry(root)
                                                           course_field = Entry(root)
                                                           sem_field = Entry(root)
                                                           form_no_field = Entry(root)
                                                           contact_no_field = Entry(root)
                                                           email_id_field = Entry(root)
                                                           address_field = Entry(root)

                                                           # bind method of widget is used for
                                                           # the binding the function with the events

                                                           # whenever the enter key is pressed
                                                           # then call the focus1 function
                                                           name_field.bind("<Return>", focus1)

                                                           # whenever the enter key is pressed
                                                           # then call the focus2 function
                                                           course_field.bind("<Return>", focus2)

                                                           # whenever the enter key is pressed
                                                           # then call the focus3 function
                                                           sem_field.bind("<Return>", focus3)

                                                           # whenever the enter key is pressed
                                                           # then call the focus4 function
                                                           form_no_field.bind("<Return>", focus4)

                                                           # whenever the enter key is pressed
                                                           # then call the focus5 function
                                                           contact_no_field.bind("<Return>", focus5)

                                                           # whenever the enter key is pressed
                                                           # then call the focus6 function
                                                           email_id_field.bind("<Return>", focus6)

                                                           # grid method is used for placing
                                                           # the widgets at respective positions
                                                           # in table like structure .
                                                           name_field.grid(row=1, column=1, ipadx="100")
                                                           course_field.grid(row=2, column=1, ipadx="100")
                                                           sem_field.grid(row=3, column=1, ipadx="100")
                                                           form_no_field.grid(row=4, column=1, ipadx="100")
                                                           contact_no_field.grid(row=5, column=1, ipadx="100")
                                                           email_id_field.grid(row=6, column=1, ipadx="100")
                                                           address_field.grid(row=7, column=1, ipadx="100")

                                                           # call excel function
                                                           excel()

                                                           # create a Submit Button and place into the root window
                                                           submit = Button(root, text="Submit", fg="Black",
                                                                           bg="Red", command=insert)
                                                           submit.grid(row=8, column=1)

                                                           # start the GUI
                                                           root.mainloop()

                                                           from random import randint

                                                           rand = randint(1, 100)
                                                           counter = 0

                                                           while True:
                                                               counter += 1
                                                               num = int(
                                                                   input("Enter values between 1 and 100 (0 exit):"))
                                                               if (num == 0):
                                                                   print("You canceled the game.")
                                                                   break
                                                               elif num < rand:
                                                                   print("Enter a Higher Number.")
                                                                   continue
                                                               elif num > rand:
                                                                   print("Enter a Lower Number.")
                                                                   continue
                                                               else:
                                                                   print("Randomly selected number {0}!".format(rand))
                                                                   print("Your guess number {0}".format(counter))

                                                                   favorite_fruits = ['apples', 'erics', 'oranges']
                                                                   print(
                                                                       "My favourite fruits {}".format(favorite_fruits))
                                                                   save
                                                                   